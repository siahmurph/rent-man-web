import streamlit as st
import pandas as pd
from google.cloud import storage
from datetime import datetime
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.pagebreak import Break

# --- 1. CONFIGURATION (Global Variables) ---
PROPERTY_NAMES = [
    "12241 Londonderry Lane",
    "1410 W. Irving Park",
    "1745 N. Clybourn Residential",
    "2949 Halsted",
    "3349 Willowcreek",
    "6854 Highland Pines Circle",
    "995 Second Ave.",
    "Eastgate Plaza",
    "Foxmoor Plaza",
    "Grand Trunk Warehouse",
    "Morton Grove 09",
    "Morton Grove 10",
    "Morton Grove 11",
    "Morton Grove 37",
    "Patriot Square",
    "Riverdale Shopping Center",
    "San Carlos Plaza",
    "Shoppes on Clybourn",
    "Willowcreek Plaza",
]

OUTPUT_ORDER = [
    "Shoppes on Clybourn",
    "Eastgate Plaza",
    "Grand Trunk Warehouse",
    "Willowcreek Plaza",
    "3349 Willowcreek",
    "995 Second Ave.",
    "Foxmoor Plaza",
    "Patriot Square",
    "Riverdale Shopping Center",
    "San Carlos Plaza",
    "1410 W. Irving Park",
    "2949 Halsted",
    "1745 N. Clybourn Residential",
    "Morton Grove 09",
    "Morton Grove 10",
    "Morton Grove 11",
    "Morton Grove 37",
    "12241 Londonderry Lane",
    "6854 Highland Pines Circle",
]

HEADER_ROW = "Account Type,Parent,Account,Amount"

# --- 2. GCS INITIALIZATION ---
# Authenticate using secrets stored in Streamlit Cloud
client = storage.Client.from_service_account_info(st.secrets["gcp"])
BUCKET_NAME = "rent-man-reports-1"


# --- 3. CORE LOGIC FUNCTIONS ---
def parse_csv_sections(content):
    """Finds the starting line index for each of the 19 property sections."""
    lines = content.strip().split("\n")
    header_indices = []
    for i, line in enumerate(lines):
        normalized_line = line.strip().replace('"', "").replace("\r", "")
        if normalized_line == HEADER_ROW:
            header_indices.append(i)
    return header_indices, lines


def transform_property_section(lines, start_idx, end_idx, property_name):
    """Processes raw lines into an aggregated expense report for a single building."""
    result_rows = [[f"Property Name: {property_name}", "", "", ""]]
    expenses, noe = [], []

    for i in range(start_idx + 1, end_idx):
        line = lines[i].strip()
        if not line:
            continue
        try:
            parts = next(csv.reader([line]))
            if len(parts) < 4:
                continue
            a_type, parent, account = (
                parts[0].strip(),
                parts[1].strip(),
                parts[2].strip(),
            )
            # Clean numeric data: remove quotes and commas
            clean_val = parts[3].replace('"', "").replace(",", "").strip()
            val = round(float(clean_val), 2)

            if a_type == "Expense":
                expenses.append([a_type, parent, account, val])
            elif a_type == "Non Operating Expense":
                noe.append([a_type, parent, account, val])
        except:
            continue

    # Group by Parent Account
    for data, label in [(expenses, "Expense"), (noe, "Non Operating Expense")]:
        grouped = {}
        for row in data:
            key = row[1] if row[1] else row[2]
            grouped[key] = round(grouped.get(key, 0) + row[3], 2)
        for key, total in sorted(grouped.items()):
            result_rows.append([label, key, "", total])

        total_sum = round(sum(r[3] for r in data), 2)
        result_rows.append(
            [f"{'NOE ' if 'Non' in label else ''}Total", "", "", total_sum]
        )

    prop_total = round(sum(r[3] for r in expenses) + sum(r[3] for r in noe), 2)
    result_rows.append(["Property Total", "", "", prop_total])
    return result_rows


def convert_df_to_excel(df):
    """Applies blue headers, borders, and page breaks for the Excel export."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")
        ws = writer.sheets["Report"]
        blue_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        total_rows = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), 1):
            val = str(row[0].value)
            for cell in row:
                cell.border = border
            if "Property Name:" in val:
                for cell in row:
                    cell.fill = blue_fill
            if "Property Total" in val:
                total_rows.append(idx)

        # Add page breaks between properties
        for row_idx in total_rows[:-1]:
            ws.row_breaks.append(Break(id=row_idx))
    return output.getvalue()


# --- 4. APP UI ---
st.set_page_config(page_title="RentManager Pro Dashboard", layout="wide")
st.title("📊 RentManager Pro Dashboard")

# Top Control Panel for better scannability
with st.container(border=True):
    col1, col2 = st.columns([1, 2])

    with col1:
        try:
            bucket = client.get_bucket(BUCKET_NAME)
            # Poll bucket for all available CSV reports
            available_files = [
                b.name for b in bucket.list_blobs() if b.name.endswith(".csv")
            ]
            selected_file = st.selectbox(
                "📂 1. Choose GCS File",
                ["Select..."] + sorted(available_files, reverse=True),
            )
        except Exception as e:
            st.error(f"GCS Error: {e}")
            selected_file = None

    with col2:
        if selected_file and selected_file != "Select...":
            with st.form("main_controls"):
                # Multi-select for filtering buildings
                selected_props = st.multiselect(
                    "🏗️ 2. Select Properties:", OUTPUT_ORDER, default=OUTPUT_ORDER
                )
                submit = st.form_submit_button(
                    "Generate & Stitch Report", use_container_width=True
                )
        else:
            submit = False

# --- 5. RESULTS GENERATION ---
if selected_file and selected_file != "Select..." and submit:
    blob = bucket.blob(selected_file)
    header_indices, lines = parse_csv_sections(blob.download_as_text())

    # Security check: Ensure file contains all 19 property headers
    if len(header_indices) == 19:
        all_rows = []
        # Process in the order specified by OUTPUT_ORDER
        for p_name in OUTPUT_ORDER:
            if p_name in selected_props:
                try:
                    # Map building name back to its position in the original file
                    orig_idx = PROPERTY_NAMES.index(p_name)
                    start_idx = header_indices[orig_idx]
                    end_idx = (
                        header_indices[orig_idx + 1]
                        if orig_idx + 1 < len(header_indices)
                        else len(lines)
                    )

                    # Stitch this section into the master list
                    all_rows.extend(
                        transform_property_section(lines, start_idx, end_idx, p_name)
                    )
                    all_rows.append(["", "", "", ""])  # Spacer row
                except Exception as e:
                    st.warning(f"Error on {p_name}: {e}")

        final_df = pd.DataFrame(
            all_rows, columns=["Account Type", "Parent", "Account", "Amount"]
        )

        # Display Results
        st.divider()
        res_col1, res_col2 = st.columns([3, 1])
        with res_col1:
            st.subheader("Report Preview")
            st.dataframe(final_df, use_container_width=True, height=600)
        with res_col2:
            st.metric("Buildings Included", len(selected_props))
            xlsx_data = convert_df_to_excel(final_df)
            st.download_button(
                "📥 Download Excel Report",
                xlsx_data,
                f"Report_{selected_file}.xlsx",
                use_container_width=True,
            )
    else:
        st.error(
            f"Validation Error: Expected 19 headers, but found {len(header_indices)} in this file."
        )

# Sidebar Clean-up
with st.sidebar:
    st.success("✅ Connected to GCS")
    st.divider()
    st.info(
        "This tool stitches raw RentManager CSV data into formatted property reports."
    )
